from openpyxl import load_workbook # 엑셀 파일 읽기 위해 사용

# 엑셀 파일의 정보 불러오기
load_wb = load_workbook(r'12. Import_Excel_Information_Automatical_Generate_Certification\수료증명단.xlsx')
load_ws = load_wb.active

# 이름, 생일, 수료증 번호를 리스트로 생성
name_list = []
birthday_list = []
ho_list = []
for i in range(1, load_ws.max_row + 1):
    name_list.append(load_ws.cell(i, 1).value)
    birthday_list.append(load_ws.cell(i, 2).value)
    ho_list.append(load_ws.cell(i, 3).value)

# 각 변수에 저장된 값 출력
print(name_list)
print(birthday_list)
print(ho_list)