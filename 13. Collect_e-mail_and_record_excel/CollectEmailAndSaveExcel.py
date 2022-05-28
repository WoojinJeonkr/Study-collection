import requests
import re
from openpyxl import Workbook, load_workbook
from openpyxl import workbook

# 접속할 뉴스 페이지 주소
url = 'https://news.v.daum.net/v/20220528172655358'

# 헤더 정보 입력
headers = {
    'User-Agent' : 'Mozilla/5.0',
    'Content-type' : 'text/html; charset=utf-8'
}

# url로 접속
response = requests.get(url, headers=headers)

# 이메일 탐색
results = re.findall(r'[\w\.-]+@[\w\.-]+', response.text)

# 탐색한 이메일에서 중복 제거
results = list(set(results))

# 출력
print(results)

try:
    wb = load_workbook(r"13. Collect_e-mail_and_record_excel\email.xlsx", data_only=True)
    sheet = wb.active
except:
    wb = Workbook()
    sheet = wb.active

for result in results:
    sheet.append([result])

wb.save(r"13. Collect_e-mail_and_record_excel\email.xlsx")